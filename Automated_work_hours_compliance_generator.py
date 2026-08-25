import os
import pandas as pd
pd.set_option('future.no_silent_downcasting', True)
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

from openpyxl import load_workbook
import numpy as np

import ast

import requests
import shutil
from zoneinfo import ZoneInfo
import logging
import sys
import re
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure the logger
logging.basicConfig(
    level=logging.DEBUG,  # Log at DEBUG level
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),  # Log to a file
        logging.StreamHandler()          # Log to console as well
    ]
)

# Redirect print to the logging module
class LoggerWriter(object):
    def __init__(self, level):
        self.level = level

    def write(self, message):
        if message.strip():  # Ignore empty messages (like newlines)
            self.level(message)

    def flush(self):
        pass

current_directory = os.getcwd()

# Redirect stdout to log to the console and the file
sys.stdout = LoggerWriter(logging.info)
sys.stderr = LoggerWriter(logging.error)
print("#############################################################################################################################################################")

today = pd.Timestamp.today()

try:
    # Works when running as a .py script
    current_directory = os.path.dirname(os.path.abspath(__file__))
except NameError:
    # Fallback for notebooks or interactive shells
    current_directory = os.getcwd()

#BASE_DIR = Path(__file__).resolve().parent

#folder_path = Path(
#    os.environ.get("FOLDER_PATH_gme_compliance", BASE_DIR)
#)

folder_path = os.environ["FOLDER_PATH_gme_compliance"]

old_file_folder = 'past_lists'

old_file_folder_path = os.path.join(folder_path, old_file_folder)

# Load Excel files into a DataFrame

folder_path_for_files = Path(folder_path)

BUSINESS_TIMEZONE = ZoneInfo("America/Los_Angeles")

def find_todays_file(filename_prefix: str) -> Path:
    date_string = datetime.now(BUSINESS_TIMEZONE).strftime("%Y%m%d")

    # Requires a two-digit suffix such as _01.csv or _02.csv.
    pattern = f"{filename_prefix}_{date_string}_[0-9][0-9][0-9][0-9].csv"
    matches = list(folder_path_for_files.glob(pattern))

    if not matches:
        raise FileNotFoundError(
            f"No file matching {pattern!r} found in {folder_path_for_files}"
        )

    # Select the newest file if multiple versions exist.
    selected_file = max(
        matches,
        key=lambda file: file.stat().st_mtime,
    )

    print(f"Using file: {selected_file}")
    return selected_file


active_trainees_file = find_todays_file(
    "CSMC_ActiveTrainees"
)

work_hours_file = find_todays_file(
    "CSMC_WorkHours"
)

active_trainees_df = pd.read_csv(active_trainees_file)
work_hours_df = pd.read_csv(work_hours_file)
#format df and drop unnecessary columns 
drop_columns =['departmentid','defaultrotationlocationid','personid']

hours = work_hours_df.drop(columns=drop_columns)
active = active_trainees_df.drop(columns=drop_columns)

#drop anyone with blank email address
active = active[active["Person's Primary E-Mail Address"].notna()]  

# generating new active file, move old into "past" folder
src_file_name = "weekly_compliance_email_list.xlsx"
src_file = os.path.join(folder_path, src_file_name)
target_folder_name = "past_lists/old_compliance_list"
target_folder = os.path.join(folder_path, target_folder_name)

os.makedirs(target_folder, exist_ok=True)


# Get today's date
today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

weekday = today.weekday()  # Monday=0, ..., Sunday=6

# Determine reference date
if weekday == 0:  # Monday
    # Get previous Thursday
    ref_date = today - timedelta(days=4)
elif weekday == 3:  # Thursday
    # Get previous Monday
    ref_date = today - timedelta(days=3)
else:
    # Default: just use today
    ref_date = today

# Format the reference date
date_str = ref_date.strftime("%m_%d_%Y")

# Build new file name
base_name, ext = os.path.splitext(os.path.basename(src_file))
new_file_name = f"{base_name}_{date_str}{ext}"
dst_file = os.path.join(target_folder, new_file_name)

# Move the file if present
if os.path.exists(src_file):
    shutil.move(src_file, dst_file)
    print(f"Moved file to: {dst_file}")
else:
    print(f"Source file does not exist: {src_file}")


hours[['Trainee Last Name', 'Trainee First Name']] = (
    hours['Person']
    .str.strip()
    .str.split(r'\s{2,}', n=1, expand=True, regex=True)
)

# Remove any leading/trailing spaces
hours['Trainee Last Name'] = hours['Trainee Last Name'].str.strip()
hours['Trainee First Name'] = hours['Trainee First Name'].str.strip()
#rename columns 


hours_columns = ["Person's National Provider Identifier", 'Person', 'Status', 'Program',
       'Work Type', 'Start Date/Time', 'End Date/Time', 'Hours Worked',
       'Rotation', 'Rotation Start Date', 'Rotation End Date', 'Source',
       'Resident Approved', 'Administrator Approved', 'Institution/Location',
       'In Violation', 'Violations', 'Rules Violated', 'Comment', 'Comment By',
       'Last Update', 'Date Logged', "Person's Coordinator Email",
       "Person's Primary E-Mail Address", "Person's Program Coordinator",
       "Person's Program Director", 'Program Director Email',
       'Trainee Last Name', 'Trainee First Name']

active_columns = ['ID Number', 'Last Name', 'First Name', 'Middle Name',
       "Person's National Provider Identifier",
       "Person's Primary E-Mail Address", 'Department/Division', 'Program',
       "Person's Program Director", 'Program Director Email', 'Status', 
       "Person's Program Start Date",
       "Person's Program End Date", "Person's Coordinator Email",
       "Person's Program Coordinator"]

#build list for email automation
table_list_columns = ["Trainee First Name", "Trainee Last Name", "Trainee Email", "Date of Missing Hours", "Week of Missing Hours", "Violations", "ResQ Violations",
                     "Program Admin First Name",	"Program Admin Last Name", "Program Admin Email", "Program Director First Name", "Program Director Last Name", "Program Director Email",
                     "80 Hr", "Day Off", "Call", "24+", "SB"]

hours_column_update = ["Person's National Provider Identifier", 'Person', 'Status', 'Program',
       'Work Type', 'Actual Start', 'Actual End', 'Actual Hours Worked', 
       'Rotation', 'Rotation Start Date', 'Rotation End Date', 'Source', 
       'Resident Approved', 'Administrator Approved', 'Institution/Location', 
       'In Violation', 'Violation(s)', 'Rules Violated', 'Comment', 'Comment By',
       'Last Update', "Date Logged", "Program Admin Email", 
       "Trainee Email", "Person's Program Coordinator", 
       "Person's Program Director", 'Program Director Email', 'Trainee Last Name', 'Trainee First Name',]

active_columns_update = ['ID Number', 'Trainee Last Name', 'Trainee First Name', 'Middle Name',
       "Person's National Provider Identifier",
       "Trainee Email", 'Department/Division', 'Program',
       "Person's Program Director", 'Program Director Email', 'Status', "Person's Program Start Date",
       "Person's Program End Date", "Program Admin Email",
       "Person's Program Coordinator"]


#update column names
hours.columns = hours_column_update
active.columns = active_columns_update

hours['Trainee Email'] = hours['Trainee Email'].str.lower()
active['Trainee Email'] = active['Trainee Email'].str.lower()

hours['Program Admin Email'] = hours['Program Admin Email'].str.lower()
active['Program Admin Email'] = active['Program Admin Email'].str.lower()

active = active[active['Status']!='Chief Resident']
# filter down hours based on week of interest
# define week of interest

# Get today's date (midnight)
today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

# Determine most recent Sunday (start of the current week)
# weekday(): Monday=0, Sunday=6 → days_since_sunday = (weekday + 1) % 7
days_since_sunday = (today.weekday() + 1) % 7
start_of_this_week = today - timedelta(days=days_since_sunday)

# Define last week's range (Sunday → next Saturday)
start_of_last_week = start_of_this_week - timedelta(days=7)
end_of_last_week = start_of_this_week - timedelta(days=1)   # Saturday

# Adjust times
start_of_last_week = start_of_last_week.replace(hour=0, minute=0, second=0, microsecond=0)  # Sunday 12:00 AM
end_of_last_week = end_of_last_week.replace(hour=23, minute=59, second=0, microsecond=0)      # Saturday 11:59 PM
actual_end_raw = (
    hours['Actual End']
    .astype('string')
    .str.strip()
)

hours['Actual End'] = pd.to_datetime(
    actual_end_raw,
    format='mixed',
    errors='coerce',
)
actual_start_raw = (
    hours['Actual Start']
    .astype('string')
    .str.strip()
)

hours['Actual Start'] = pd.to_datetime(
    actual_start_raw,
    format='mixed',
    errors='coerce',
)

mask = (
    (hours['Actual End'] >= start_of_last_week) &
    (hours['Actual Start'] <= end_of_last_week)
)

df_last_week = hours.loc[mask].copy()

# Filter rows between end and start of week
#mask = (hours['Actual End'] >= start_of_last_week) & (hours['Actual Start'] <= end_of_last_week)
#df_last_week = hours.loc[mask].copy() 

#use different week for resQ and violations, we don't need to check for cross over times
mask_non_hours = (hours['Actual Start'] >= start_of_last_week) & (hours['Actual Start'] <= end_of_last_week)
df_last_week_non_hours = hours.loc[mask_non_hours].copy() 


resQ = df_last_week_non_hours[df_last_week_non_hours['Work Type']=='ResQ Working']

df_last_week_non_hours['In Violation'] = df_last_week_non_hours['In Violation'].str.strip().str.lower()

valid_yes = ['yes', 'y']

violations = df_last_week_non_hours[df_last_week_non_hours['In Violation'].isin(valid_yes)]

#missing_hours
unique_emails_hours_entry = df_last_week["Trainee Email"].unique().tolist()


# Convert to set for faster lookup
email_set_hours = set(unique_emails_hours_entry)

# Get emails in df1 that are NOT in df2
emails_not_in_hours = set(active["Trainee Email"]) - email_set_hours

# Convert to a list if you want
emails_not_in_hours = list(emails_not_in_hours)

print(emails_not_in_hours)


violations['Violations'] = violations['Actual Start'].dt.strftime('%m/%d/%Y') +' '+ violations['Rules Violated']
# group by unique email
consolidated_violations = (
    violations.groupby(['Trainee Email'], as_index=False)
      .agg({
          'Trainee First Name': 'first',
          'Trainee Last Name': 'first',
          'Program Admin Email':'first',
          'Program':'first',
          'Violations': lambda x: ', '.join(sorted(set(x.dropna())))
      })
)
consolidated_resQ = (
    resQ.groupby(['Trainee Email'], as_index=False)
      .agg({
          'Trainee First Name': 'first',
          'Trainee Last Name': 'first',
          'Program Admin Email':'first',
          'Program':'first',
          'Work Type': lambda x: ', '.join(sorted(set(x.dropna())))
      })
)
consolidated_resQ['ResQ Violations'] = 'Yes'
consolidated_resQ = consolidated_resQ.drop('Work Type', axis=1)
# partial hour inclusion to the missing hours variable 

#get count of days each day has added hours, if that count is less than 5, they are considered missing hours
df_last_week['Shift Date'] = df_last_week['Actual Start'].dt.date
days_worked = (
    df_last_week.groupby('Trainee Email')['Actual Start']
      .nunique()   # count unique dates
      .reset_index(name='Days Worked')
)

less_than_4 = days_worked[days_worked['Days Worked'] < 4]

df_filtered = df_last_week[df_last_week['Trainee Email'].isin(less_than_4['Trainee Email'])]

df_filtered = df_filtered.drop(columns= 'Shift Date')

df_filtered_unique = df_filtered.drop_duplicates(subset='Trainee Email', keep='first')

def expand_shift_days(row):
    start = row['Actual Start'].normalize()
    end = row['Actual End'].normalize()

    # Generate daily date range
    return pd.date_range(start, end, freq='D').date




df_last_week['Days Covered'] = df_last_week.apply(expand_shift_days, axis=1)
df_days_all = df_last_week.explode('Days Covered')

#need to remove days that are being covered outside of week due to crossover shifts
df_days = df_days_all[df_days_all['Days Covered']>= start_of_last_week.date()]

days_worked = (
    df_days.groupby('Trainee Email')['Days Covered']
           .nunique()
           .reset_index(name='Days Worked')
)

less_than_4 = days_worked[days_worked['Days Worked'] < 4]

df_filtered = df_last_week[
    df_last_week['Trainee Email'].isin(less_than_4['Trainee Email'])
]

df_filtered_unique = df_filtered.drop_duplicates(
    subset='Trainee Email',
    keep='first'
)

consolidated_partial_hours_miss = (
    df_filtered_unique.groupby(['Trainee Email'], as_index=False)
      .agg({
          'Trainee First Name': 'first',
          'Trainee Last Name': 'first',
          'Program':'first',
          'Program Admin Email':'first'
      })
)

consolidated_partial_hours_miss['Week of Missing Hours'] = start_of_last_week.strftime('%m/%d/%Y') +'-' + end_of_last_week.strftime('%m/%d/%Y')
#get hours together
hours_miss = active[active['Trainee Email'].isin(emails_not_in_hours)]
consolidated_hours_miss = (
    hours_miss.groupby(['Trainee Email'], as_index=False)
      .agg({
          'Trainee First Name': 'first',
          'Trainee Last Name': 'first',
          'Program':'first',
          'Program Admin Email':'first'
      })
)
consolidated_hours_miss['Week of Missing Hours'] = start_of_last_week.strftime('%m/%d/%Y') +'-' + end_of_last_week.strftime('%m/%d/%Y')
total_consolidated_hours_miss = pd.concat([consolidated_hours_miss,consolidated_partial_hours_miss], ignore_index= True)
total_consolidated_hours_miss_unique = total_consolidated_hours_miss.drop_duplicates(subset='Trainee Email', keep='first')



#join together 
table = pd.concat([consolidated_resQ, total_consolidated_hours_miss_unique, consolidated_violations], ignore_index= True)
id_cols = ['Trainee Email']

# All other columns to keep, filling NaNs where possible
value_cols = [col for col in table.columns if col not in id_cols]

# Group by email and take the first non-NaN for each column
consolidated_df = table.groupby(id_cols, as_index=False).agg(
    {col: 'first' for col in value_cols}
)


# update directors name
active[['Program Director First Name', 'Program Director Last Name']] = (
    active["Person's Program Director"]
    .astype('string')
    .str.strip()
    .str.split(r'\s+', n=1, expand=True, regex=True)
)
active_columns_to_use = [
    'Trainee Email',
    'Program Director First Name',
    'Program Director Last Name',
    'Program Director Email',
]

# Normalize the join key.
for df in [active, consolidated_df]:
    df['Trainee Email'] = (
        df['Trainee Email']
        .astype('string')
        .str.strip()
        .str.lower()
    )

active_lookup = (
    active[active_columns_to_use]
    .drop_duplicates()
)

consolidated_df = consolidated_df.merge(
    active_lookup,
    how='left',
    on='Trainee Email',
    validate='many_to_one',
    indicator='_active_match',
)

# Confirm every consolidated trainee was found in active.
missing_emails = (
    consolidated_df.loc[
        consolidated_df['_active_match'].eq('left_only'),
        'Trainee Email',
    ]
    .dropna()
    .drop_duplicates()
    .tolist()
)

consolidated_df = consolidated_df.drop(columns='_active_match')
# remove test cases from jeffrey.mckelvey@cshs.org	
consolidated_df = consolidated_df[consolidated_df['Trainee Email']!='jeffrey.mckelvey@cshs.org']
# Added filter for Pilot Programs
pilots = ['NEUROSURG-Neurological Surgery-ACGME', 'Imaging-Diagnostic Radiology-ACGME', 'MED-Pulmonary Disease & Critical Care Medicine-ACGME',
          'RAD-Radiation Oncology-ACGME', 'PEDS-Pediatric Medicine-ACGME', 'Surgery-Advanced GI MIS/Bariatric', 'MED-Hospice & Palliative Care Medicine-ACGME',
          'OB/GYN-Obstetrics & Gynecology-ACGME', 'MED-Rheumatology-ACGME', 'MED-Rheumatology-Research']
consolidated_df = consolidated_df[consolidated_df['Program'].isin(pilots)]

# Load Excel files into a DataFrame

active_file_name = 'active.csv'
hours_file_name = 'hours.csv'

new_active_file_path = os.path.join(old_file_folder_path, 'old_active_list')
new_hours_file_path = os.path.join(old_file_folder_path, 'old_hours_list')

date_str = datetime.today().strftime("%m_%d_%Y")  # e.g., "10_22_2025"

# Build new file name
base_name, ext = os.path.splitext(os.path.basename(active_file_name))
new_active_file_name = f"{base_name}_{date_str}{ext}"

# Get the full destination path
active_destination_file = os.path.join(new_active_file_path, new_active_file_name)

# Move the file
shutil.move(active_trainees_file, active_destination_file)
#print(source_file, destination_file)
print(f'Moved: {new_active_file_name}')

# Build new file name
base_name, ext = os.path.splitext(os.path.basename(hours_file_name))
new_hours_file_name = f"{base_name}_{date_str}{ext}"

# Get the full destination path
hours_destination_file = os.path.join(new_hours_file_path, new_hours_file_name)


shutil.move(work_hours_file, hours_destination_file)
#print(source_file, destination_file)
print(f'Moved: {new_hours_file_name}')
qc_df = (consolidated_df
        .groupby('Program')
        .size()
        .reset_index(name='Count')
        )
qc_df['SummaryLine'] = qc_df['Program'] + ' → ' + qc_df['Count'].astype(str) + ' trainees'

# Convert 'SummaryLine' column into a single string with newline separators
full_summary = '\n'.join(qc_df['SummaryLine'].astype(str))

# Optionally, save it back to Excel in a new column or a single cell
# For example, create a new row with the concatenated summary
summary_df = pd.DataFrame({'FullSummary': [full_summary]})


compliance_list_name = 'weekly_compliance_email_list.xlsx'
compliance_list_location = os.path.join(folder_path, compliance_list_name)

# --- Save both DataFrames ---
with pd.ExcelWriter(compliance_list_location, engine='openpyxl') as writer:
    consolidated_df.to_excel(writer, sheet_name="Sheet1", index=False)
    summary_df.to_excel(writer, sheet_name="Sheet2", index=False)

# ensure writer handle released
try:
    del writer
except NameError:
    pass


import time
# small pause to let OS release file lock
time.sleep(2)   # try 2 seconds; increase to 3-4 if still locked

import gc
# force garbage collection
gc.collect()

# --- Reopen workbook to add tables ---
wb = load_workbook(compliance_list_location)

# --- Table for Sheet1 ---
ws1 = wb["Sheet1"]
num_rows1 = consolidated_df.shape[0] + 1
num_cols1 = consolidated_df.shape[1]
table_ref1 = f"A1:{chr(64 + num_cols1)}{num_rows1}"
tab1 = Table(displayName="Table1", ref=table_ref1)
style1 = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
tab1.tableStyleInfo = style1
ws1.add_table(tab1)

# --- Table for Sheet2 ---
ws2 = wb["Sheet2"]
num_rows2 = summary_df.shape[0] + 1
num_cols2 = summary_df.shape[1]
table_ref2 = f"A1:{chr(64 + num_cols2)}{num_rows2}"
tab2 = Table(displayName="Table2", ref=table_ref2)
style2 = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
tab2.tableStyleInfo = style2
ws2.add_table(tab2)

# --- Save ---
wb.save(compliance_list_location)
wb.close()
del wb
gc.collect()
import os
try:
    with open(compliance_list_location, 'a'):
        print("File is writable")
except Exception as e:
    print("File still locked:", e)

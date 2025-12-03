# optimized_monthly_compliance.py
import os
import shutil
import logging
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import gc
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------- CONFIG ----------
folder_path = os.environ["FOLDER_PATH_gme_compliance"]
old_file_folder = 'past_lists'
old_file_folder_path = os.path.join(folder_path, old_file_folder)

PILOT_ONLY = True
PILOTS = ['NEUROSURG-Neurological Surgery-ACGME', 'Imaging-Diagnostic Radiology-ACGME']
OUTPUT_PREFIX = "monthly_compliance_email_list"

HOURS_COLS_NEW = ["Person's National Provider Identifier", 'Person', 'Status', 'Program',
       'Work Type', 'Actual Start', 'Actual End', 'Actual Hours Worked',
       'Rotation', 'Rotation Start Date', 'Rotation End Date', 'Source',
       'Resident Approved', 'Administrator Approved', 'Institution/Location',
       'In Violation', 'Violation(s)', 'Rules Violated', 'Comment', 'Comment By',
       'Last Update', "Date Logged", "Program Admin Email",
       "Trainee Email", "Person's Program Coordinator",
       "Person's Program Director", 'Trainee Last Name', 'Trainee First Name']

ACTIVE_COLS_NEW = ['ID Number', 'Trainee Last Name', 'Trainee First Name', 'Middle Name',
       "Person's National Provider Identifier",
       "Trainee Email", 'Department/Division', 'Program',
       "Person's Program Director", 'Status', "Person's Program Start Date",
       "Person's Program End Date", "Program Admin Email",
       "Person's Program Coordinator"]

PD_LIST_COLS_NEW = ['Program', 'programtype', 'department', 'Program Director First Name',
       'Program Director Last Name', 'programdirector', 'Program Director Email',
       'programcoordinator', 'Program Admin Email']

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')

# ---------- Utilities ----------
def ensure_dirs():
    os.makedirs(old_file_folder_path, exist_ok=True)
    os.makedirs(os.path.join(old_file_folder_path, 'old_active_list'), exist_ok=True)
    os.makedirs(os.path.join(old_file_folder_path, 'old_hours_list'), exist_ok=True)
    os.makedirs(os.path.join(folder_path, 'past_lists', 'old_compliance_list'), exist_ok=True)

def read_inputs():
    active = pd.read_excel(os.path.join(folder_path, 'active.xlsx'))
    hours = pd.read_excel(os.path.join(folder_path, 'hours.xlsx'))
    pd_list = pd.read_excel(os.path.join(folder_path, 'PD_and_PA_report_list.xlsx'))
    return active, hours, pd_list

def normalize_and_clean(active, hours, pd_list):
    # Split 'Person' into first/last names for hours
    hours[['Trainee Last Name', 'Trainee First Name']] = hours['Person'].str.split(',', n=1, expand=True)
    hours['Trainee Last Name'] = hours['Trainee Last Name'].str.strip()
    hours['Trainee First Name'] = hours['Trainee First Name'].str.strip()
    
    # Rename columns
    hours.columns = HOURS_COLS_NEW
    active.columns = ACTIVE_COLS_NEW
    pd_list.columns = PD_LIST_COLS_NEW

    # Lowercase emails
    for df, col in [(hours, 'Trainee Email'), (active, 'Trainee Email'),
                    (hours, 'Program Admin Email'), (active, 'Program Admin Email'),
                    (pd_list, 'Program Admin Email')]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lower().replace({'nan': np.nan})

    # Remove Chief Residents
    if 'Status' in active.columns:
        active = active[active['Status'] != 'Chief Resident']

    # Parse datetimes
    for c in ['Actual Start', 'Actual End', 'Date Logged', 'Last Update']:
        if c in hours.columns:
            hours[c] = pd.to_datetime(hours[c], errors='coerce')
    return active, hours, pd_list

def prev_month_range(reference_date=None):
    if reference_date is None:
        reference_date = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    first_of_this_month = reference_date.replace(day=1)
    end_last_month = first_of_this_month - timedelta(days=1)
    start_last_month = end_last_month.replace(day=1)
    return start_last_month, end_last_month


def generate_full_weeks_for_month(start_month, end_month):
    """
    Generates weekly periods for the month of interest.
    Keeps weeks that END inside the month.
    Skips weeks that end in the next month.
    """

    # Example: start_month = datetime(2025, 11, 1)
    # end_month   = datetime(2025, 11, 30)

    month_start = start_month.replace(day=1)
    month_end = end_month

    # Find first Sunday on or before month_start
    first_sunday = month_start - timedelta(days=(month_start.weekday() + 1) % 7)

    weeks = []
    current_start = first_sunday

    while current_start <= month_end:
        current_end = current_start + timedelta(days=6)

        # --- KEY RULE: Skip if week ends outside target month ---
        if current_end.month != start_month.month:
            break

        # Format week label
        week_label = f"{current_start.strftime('%Y-%m-%d')} to {current_end.strftime('%Y-%m-%d')}"
        weeks.append((current_start, current_end, week_label))

        current_start += timedelta(days=7)

    return weeks


# ---------- Core Processing ----------
def process_month(active, hours, pd_list, start_month, end_month):
    active = active.copy()
    active['Trainee Email'] = active['Trainee Email'].str.lower().str.strip()
    hours['Trainee Email'] = hours['Trainee Email'].str.lower().str.strip()
    active_emails = set(active['Trainee Email'].dropna())

    trainee_info = {}
    violations_map = {}
    resq_map = {}
    missing_weeks_map = {}

    # Pre-fill trainee info from active
    for _, row in active.iterrows():
        email = row.get('Trainee Email')
        if pd.isna(email):
            continue
        trainee_info[email] = {
            'Trainee First Name': row.get('Trainee First Name'),
            'Trainee Last Name': row.get('Trainee Last Name'),
            'Program': row.get('Program'),
            'Program Admin Email': row.get('Program Admin Email')
        }

    weeks = generate_full_weeks_for_month(start_month, end_month)

    for ws, we, week_label in weeks:
        mask = (hours['Actual Start'] >= ws) & (hours['Actual Start'] <= we)
        hours_week = hours.loc[mask].copy()

        # RESQ detection
        resq_entries = hours_week[hours_week['Work Type'].str.contains('ResQ', na=False, case=False)]
        for _, r in resq_entries.iterrows():
            email = r.get('Trainee Email')
            if pd.isna(email): continue
            resq_map[email] = True
            if email not in trainee_info:
                trainee_info[email] = {
                    'Trainee First Name': r.get('Trainee First Name'),
                    'Trainee Last Name': r.get('Trainee Last Name'),
                    'Program': r.get('Program'),
                    'Program Admin Email': r.get('Program Admin Email')
                }

        # Violations detection
        if 'In Violation' in hours_week.columns:
            inv_series = hours_week['In Violation'].astype(str).str.strip().str.lower()
            valid_yes = inv_series.isin(['yes','y'])
            violations_entries = hours_week.loc[valid_yes]
            for _, v in violations_entries.iterrows():
                email = v.get('Trainee Email')
                if pd.isna(email): continue
                viol_msg = f"{v.get('Actual Start').strftime('%m/%d/%Y') if pd.notna(v.get('Actual Start')) else ''} {v.get('Rules Violated','')}"
                violations_map.setdefault(email, set()).add(viol_msg.strip())
                if email not in trainee_info:
                    trainee_info[email] = {
                        'Trainee First Name': v.get('Trainee First Name'),
                        'Trainee Last Name': v.get('Trainee Last Name'),
                        'Program': v.get('Program'),
                        'Program Admin Email': v.get('Program Admin Email')
                    }

        # Missing hours: no entries
        emails_this_week = set(hours_week['Trainee Email'].dropna())
        no_entry_emails = active_emails - emails_this_week
        for email in no_entry_emails:
            missing_weeks_map.setdefault(email, set()).add(week_label)

        # Partial coverage (<5 days)
        if not hours_week.empty:
            def expand_shift_days(row):
                s, e = row['Actual Start'], row['Actual End']
                if pd.isna(s) or pd.isna(e): return []
                start_date, end_date = s.date(), e.date()
                if end_date < start_date: return [start_date]
                return list(pd.date_range(start_date, end_date).date)
            hours_week['Days Covered'] = hours_week.apply(expand_shift_days, axis=1)
            df_days = hours_week.explode('Days Covered')
            days_worked = df_days.groupby('Trainee Email')['Days Covered'].nunique().reset_index()
            partials = days_worked[days_worked['Days Covered'] < 5]
            for _, p in partials.iterrows():
                email = p['Trainee Email']
                missing_weeks_map.setdefault(email, set()).add(week_label)

    # Build final DataFrame — only include trainees who have at least one issue
    all_emails = set(trainee_info.keys()) | set(violations_map.keys()) | set(resq_map.keys()) | set(missing_weeks_map.keys())

    # Keep only emails that actually have a problem
    def has_issue(email):
        if email in resq_map and resq_map.get(email, False):
            return True
        if email in violations_map and violations_map.get(email):
            return True
        if email in missing_weeks_map and missing_weeks_map.get(email):
            return True
        return False

    filtered_emails = {e for e in all_emails if e is not None and not pd.isna(e) and has_issue(e)}

    rows = []
    for email in sorted(filtered_emails):
        info = trainee_info.get(email, {})
        rows.append({
            'Trainee Email': email,
            'Trainee First Name': info.get('Trainee First Name'),
            'Trainee Last Name': info.get('Trainee Last Name'),
            'Program': info.get('Program'),
            'Program Admin Email': info.get('Program Admin Email'),
            'ResQ Violations': 'Yes' if resq_map.get(email, False) else np.nan,
            'Violations': ', '.join(sorted(violations_map.get(email, []))) if violations_map.get(email) else np.nan,
            'Week(s) of Missing Hours': ', '.join(sorted(missing_weeks_map.get(email, []))) if missing_weeks_map.get(email) else np.nan
        })

    consolidated_df = pd.DataFrame(rows)

    # Optional pilot filter
    if PILOT_ONLY:
        consolidated_df = consolidated_df[consolidated_df['Program'].isin(PILOTS)]

    return consolidated_df

# ---------- Output & Save ----------
def save_output(consolidated_df, start_month, end_month, program_counts_df):
    month_label_short = start_month.strftime("%m_%Y")
    output_name = f"{OUTPUT_PREFIX}_{month_label_short}.xlsx"
    out_path = os.path.join(folder_path, output_name)

    # QC summary
    qc_df = consolidated_df.groupby('Program').size().reset_index(name='Count')
    qc_df['SummaryLine'] = qc_df['Program'] + ' → ' + qc_df['Count'].astype(str) + ' trainees'
    full_summary = '\n'.join(qc_df['SummaryLine'].astype(str))
    summary_df = pd.DataFrame({'FullSummary': [full_summary]})

    # Save Excel
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        consolidated_df.to_excel(writer, sheet_name="Sheet1", index=False)
        summary_df.to_excel(writer, sheet_name="Sheet2", index=False)
        program_counts_df.to_excel(writer, sheet_name='Program Counts', index=False)


    wb = load_workbook(out_path)
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]

    # Add tables
    num_rows1, num_cols1 = consolidated_df.shape[0] + 1, consolidated_df.shape[1] if consolidated_df.shape[1] > 0 else 1
    tab1 = Table(displayName="Table1", ref=f"A1:{chr(64+num_cols1)}{num_rows1}")
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws1.add_table(tab1)

    num_rows2, num_cols2 = summary_df.shape[0] + 1, summary_df.shape[1]
    tab2 = Table(displayName="Table2", ref=f"A1:{chr(64+num_cols2)}{num_rows2}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws2.add_table(tab2)

    wb.save(out_path)
    wb.close()
    logging.info(f"Saved output to {out_path}")
    return out_path

# ---------- Output & Save ----------
def save_output(consolidated_df, start_month, end_month, program_counts_df, folder_path, OUTPUT_PREFIX):
    month_label_short = start_month.strftime("%m_%Y")
    output_name = f"{OUTPUT_PREFIX}_{month_label_short}.xlsx"
    out_path = os.path.join(folder_path, output_name)

    # QC summary
    qc_df = consolidated_df.groupby('Program').size().reset_index(name='Count')
    qc_df['SummaryLine'] = qc_df['Program'] + ' → ' + qc_df['Count'].astype(str) + ' trainees'
    full_summary = '\n'.join(qc_df['SummaryLine'].astype(str))
    summary_df = pd.DataFrame({'FullSummary': [full_summary]})

    # --- Save Excel using pandas.ExcelWriter ---
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        consolidated_df.to_excel(writer, sheet_name="Sheet1", index=False)
        summary_df.to_excel(writer, sheet_name="Sheet2", index=False)
        program_counts_df.to_excel(writer, sheet_name='Program Counts', index=False)

    # --- Reopen with openpyxl to add tables ---
    wb = load_workbook(out_path)

    # Sheet1 table
    ws1 = wb["Sheet1"]
    num_rows1 = consolidated_df.shape[0] + 1
    num_cols1 = consolidated_df.shape[1] if consolidated_df.shape[1] > 0 else 1
    tab1 = Table(displayName="Table1", ref=f"A1:{chr(64+num_cols1)}{num_rows1}")
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws1.add_table(tab1)

    # Sheet2 table
    ws2 = wb["Sheet2"]
    num_rows2, num_cols2 = summary_df.shape[0] + 1, summary_df.shape[1]
    tab2 = Table(displayName="Table2", ref=f"A1:{chr(64+num_cols2)}{num_rows2}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws2.add_table(tab2)

    # Program Counts table
    ws3 = wb["Program Counts"]
    num_rows3, num_cols3 = program_counts_df.shape[0] + 1, program_counts_df.shape[1]
    tab3 = Table(displayName="Table3", ref=f"A1:{chr(64+num_cols3)}{num_rows3}")
    tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws3.add_table(tab3)

    # Save and close workbook
    wb.save(out_path)
    wb.close()
    del wb
    gc.collect()  # release file handles

    # Optional: test file is writable
    try:
        with open(out_path, 'a'):
            logging.info(f"File is writable and ready for sync: {out_path}")
    except Exception as e:
        logging.warning(f"File still locked: {e}")

    logging.info(f"Saved output to {out_path}")
    return out_path

def archive_inputs():
    date_str = datetime.today().strftime("%m_%d_%Y")
    active_file = os.path.join(folder_path, 'active.xlsx')
    hours_file = os.path.join(folder_path, 'hours.xlsx')

    active_dest = os.path.join(old_file_folder_path, 'old_active_list', f"active_{date_str}.xlsx")
    hours_dest = os.path.join(old_file_folder_path, 'old_hours_list', f"hours_{date_str}.xlsx")

    for src, dst in [(active_file, active_dest), (hours_file, hours_dest)]:
        if os.path.exists(src):
            shutil.move(src, dst)
            logging.info(f"Moved {src} -> {dst}")
        else:
            logging.warning(f"File not found, not moved: {src}")


# ---------- Main ----------
def main():
    logging.info("Starting monthly compliance processing...")
    ensure_dirs()
    active, hours, pd_list = read_inputs()
    active, hours, pd_list = normalize_and_clean(active, hours, pd_list)

    start_month, end_month = prev_month_range()
    logging.info(f"Analyzing previous month: {start_month.date()} -> {end_month.date()}")

    # ---------- 1. Create consolidated_df ----------
    consolidated_df = process_month(active, hours, pd_list, start_month, end_month)

    # ---------- 2. Clean Program column for merging ----------
    consolidated_df['Program'] = consolidated_df['Program'].astype(str).str.strip()
    programs_info = pd_list.copy()
    programs_info['Program'] = programs_info['Program'].astype(str).str.strip()

    # ---------- 3. Prepare program/director/admin info from pd_list ----------
    # Ensure all needed columns exist
    for col in ['Program Director First Name', 'Program Director Last Name', 'Program Director Email',
                'programcoordinator', 'Program Admin Email']:
        if col not in programs_info.columns:
            programs_info[col] = np.nan

    # Split 'programcoordinator' into First/Last
    programs_info['programcoordinator'] = programs_info['programcoordinator'].fillna('')
    def split_coordinator(x):
        parts = x.split(',', 1)
        if len(parts) == 2:
            last, first = parts
        else:
            last = parts[0]
            first = ''
        return pd.Series([last.strip(), first.strip()])

    programs_info[['Program Admin Last Name', 'Program Admin First Name']] = programs_info['programcoordinator'].apply(split_coordinator)


    # Keep only needed columns
    pd_info = programs_info[['Program', 'Program Director First Name', 'Program Director Last Name',
                            'Program Director Email', 'Program Admin First Name', 'Program Admin Last Name',
                            'Program Admin Email']]

    # Deduplicate one row per Program
    pd_info = pd_info.drop_duplicates(subset=['Program'])

    # ---------- 4. Merge info into consolidated_df (Sheet1) ----------
    # Clean Program columns for merge
    consolidated_df['Program_clean'] = consolidated_df['Program'].str.lower().str.strip()
    pd_info['Program_clean'] = pd_info['Program'].str.lower().str.strip()


    consolidated_df = consolidated_df.merge(pd_info, on='Program_clean', how='left', suffixes=('', '_info'))


    consolidated_df = consolidated_df.drop(columns=['Program_clean'])

    # ---------- 5. Build program_counts_df (Sheet2) ----------
    all_programs = sorted(consolidated_df['Program'].astype(str).str.strip().unique())
    program_counts = consolidated_df['Program'].value_counts().reindex(all_programs, fill_value=0)

    program_counts_df = pd.DataFrame({
        'Program': program_counts.index,
        'Count': program_counts.values
    })

    # Merge program/director/admin info into program_counts_df
    program_counts_df['Program_clean'] = program_counts_df['Program'].str.lower().str.strip()
    program_counts_df = program_counts_df.merge(pd_info, on='Program_clean', how='left')
    program_counts_df = program_counts_df.drop(columns=['Program_clean'])

    save_output(consolidated_df, start_month, end_month, program_counts_df, folder_path, OUTPUT_PREFIX)
    archive_inputs()
    logging.info("Processing complete.")

if __name__ == "__main__":
    main()
